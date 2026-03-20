#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
回测结果分析 & Excel报告生成
============================
读取 batch_backtest_all_boards.py 的CSV输出，生成专业Excel报告

使用方法：
    python generate_report.py

依赖：openpyxl, pandas
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    exit(1)

OUTPUT_DIR = Path('backtest_results')
CAPTURE_THRESHOLD = 0.60

# 样式定义
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
NEUTRAL_FILL = PatternFill('solid', fgColor='FFEB9C')
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=12, color='2E75B6')
NUM_FONT = Font(name='Consolas', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(cell, is_pct=False, is_ratio=False):
    cell.font = NUM_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if is_pct and isinstance(cell.value, (int, float)):
        cell.number_format = '+0.0%;-0.0%'
        cell.value = cell.value / 100
    if is_ratio and isinstance(cell.value, (int, float)):
        cell.number_format = '0.0%'


def create_summary_sheet(wb, df):
    """创建总览页"""
    ws = wb.active
    ws.title = '总览'
    ws.sheet_properties.tabColor = '1F4E79'
    
    # 标题
    ws['A1'] = '全板块回测结果总览'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f'回测区间: 2024-01-01 ~ 2025-12-31 | 合格线: 捕获率 >= {CAPTURE_THRESHOLD:.0%} | 生成: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(name='Arial', size=9, color='666666')
    ws.merge_cells('A2:H2')
    
    # 板块汇总表
    row = 4
    ws.cell(row=row, column=1, value='一、板块汇总').font = SUBTITLE_FONT
    row += 1
    
    headers = ['板块', '标的数', '有信号', '合格数', '合格率', '平均策略收益', '平均股价涨幅', '判定']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    
    valid = df[df['status'] == 'success']
    
    for board in ['创业板', '科创板', '主板']:
        bdf = valid[valid['board'] == board]
        if len(bdf) == 0:
            continue
        bp = bdf[bdf['passed']]
        rate = len(bp) / len(bdf) if len(bdf) > 0 else 0
        avg_ret = bdf['strategy_return_pct'].mean()
        avg_price = bdf['price_change_pct'].mean()
        verdict = '达标' if rate >= 0.6 else '未达标'
        
        ws.cell(row=row, column=1, value=board).border = THIN_BORDER
        ws.cell(row=row, column=2, value=len(df[df['board']==board])).border = THIN_BORDER
        ws.cell(row=row, column=3, value=len(bdf)).border = THIN_BORDER
        ws.cell(row=row, column=4, value=len(bp)).border = THIN_BORDER
        
        c5 = ws.cell(row=row, column=5, value=rate)
        c5.number_format = '0.0%'
        c5.fill = PASS_FILL if rate >= 0.6 else FAIL_FILL
        c5.border = THIN_BORDER
        
        c6 = ws.cell(row=row, column=6, value=avg_ret/100)
        c6.number_format = '+0.0%;-0.0%'
        c6.border = THIN_BORDER
        
        c7 = ws.cell(row=row, column=7, value=avg_price/100)
        c7.number_format = '+0.0%;-0.0%'
        c7.border = THIN_BORDER
        
        c8 = ws.cell(row=row, column=8, value=verdict)
        c8.fill = PASS_FILL if rate >= 0.6 else FAIL_FILL
        c8.font = Font(bold=True, color='006100' if rate >= 0.6 else '9C0006')
        c8.border = THIN_BORDER
        
        row += 1
    
    # 总计行
    total_valid = len(valid)
    total_passed = len(valid[valid['passed']])
    overall_rate = total_passed / total_valid if total_valid > 0 else 0
    
    ws.cell(row=row, column=1, value='合计').font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=len(df)).font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3, value=total_valid).font = Font(bold=True)
    ws.cell(row=row, column=3).border = THIN_BORDER
    ws.cell(row=row, column=4, value=total_passed).font = Font(bold=True)
    ws.cell(row=row, column=4).border = THIN_BORDER
    
    c5 = ws.cell(row=row, column=5, value=overall_rate)
    c5.number_format = '0.0%'
    c5.font = Font(bold=True, size=12)
    c5.fill = PASS_FILL if overall_rate >= 0.6 else FAIL_FILL
    c5.border = THIN_BORDER
    
    c6 = ws.cell(row=row, column=6, value=valid['strategy_return_pct'].mean()/100 if total_valid > 0 else 0)
    c6.number_format = '+0.0%;-0.0%'
    c6.font = Font(bold=True)
    c6.border = THIN_BORDER
    
    c7 = ws.cell(row=row, column=7, value=valid['price_change_pct'].mean()/100 if total_valid > 0 else 0)
    c7.number_format = '+0.0%;-0.0%'
    c7.font = Font(bold=True)
    c7.border = THIN_BORDER
    
    overall_verdict = '策略有效' if overall_rate >= 0.6 else '需要调整'
    c8 = ws.cell(row=row, column=8, value=overall_verdict)
    c8.fill = PASS_FILL if overall_rate >= 0.6 else FAIL_FILL
    c8.font = Font(bold=True, size=12, color='006100' if overall_rate >= 0.6 else '9C0006')
    c8.border = THIN_BORDER
    
    row += 2
    
    # 场景分析
    ws.cell(row=row, column=1, value='二、场景分析').font = SUBTITLE_FONT
    row += 1
    
    headers2 = ['场景', '标的数', '合格数', '合格率', '平均策略收益', '说明']
    for c, h in enumerate(headers2, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers2))
    row += 1
    
    scenarios = {
        '上涨+盈利': '核心场景：策略捕获上涨趋势',
        '上涨+亏损': '问题场景：股价涨了但策略亏了',
        '下跌+盈利': '优秀场景：逆势盈利',
        '下跌+亏损': '风控场景：看亏损是否可控',
        '股价持平': '看策略能否在震荡中获利',
    }
    
    for scenario, desc in scenarios.items():
        subset = valid[valid['scenario'] == scenario]
        if len(subset) == 0:
            continue
        sp = len(subset[subset['passed']])
        rate = sp / len(subset) if len(subset) > 0 else 0
        avg_ret = subset['strategy_return_pct'].mean()
        
        ws.cell(row=row, column=1, value=scenario).border = THIN_BORDER
        ws.cell(row=row, column=2, value=len(subset)).border = THIN_BORDER
        ws.cell(row=row, column=3, value=sp).border = THIN_BORDER
        c4 = ws.cell(row=row, column=4, value=rate)
        c4.number_format = '0.0%'
        c4.border = THIN_BORDER
        c5 = ws.cell(row=row, column=5, value=avg_ret/100)
        c5.number_format = '+0.0%;-0.0%'
        c5.border = THIN_BORDER
        ws.cell(row=row, column=6, value=desc).border = THIN_BORDER
        row += 1
    
    # 列宽
    widths = [10, 8, 8, 8, 10, 14, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_detail_sheet(wb, df, board_name, tab_color):
    """创建板块明细页"""
    valid = df[(df['board'] == board_name) & (df['status'] == 'success')]
    if len(valid) == 0:
        return
    
    ws = wb.create_sheet(title=board_name)
    ws.sheet_properties.tabColor = tab_color
    
    ws['A1'] = f'{board_name} 回测明细'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:L1')
    
    row = 3
    headers = ['代码', '名称', '行业', '股价涨幅', '最大涨幅', '策略收益', 
               '捕获率', '场景', '交易数', '胜率', '最大盈利', '合格']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    
    # 按策略收益排序
    valid_sorted = valid.sort_values('strategy_return_pct', ascending=False)
    
    for _, r in valid_sorted.iterrows():
        ws.cell(row=row, column=1, value=r['ts_code']).border = THIN_BORDER
        ws.cell(row=row, column=2, value=r['name']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=r.get('industry', '')).border = THIN_BORDER
        
        c4 = ws.cell(row=row, column=4, value=r['price_change_pct']/100)
        c4.number_format = '+0.0%;-0.0%'
        c4.border = THIN_BORDER
        
        c5 = ws.cell(row=row, column=5, value=r['max_gain_pct']/100)
        c5.number_format = '+0.0%;-0.0%'
        c5.border = THIN_BORDER
        
        c6 = ws.cell(row=row, column=6, value=r['strategy_return_pct']/100)
        c6.number_format = '+0.0%;-0.0%'
        c6.font = Font(color='006100' if r['strategy_return_pct'] > 0 else '9C0006')
        c6.border = THIN_BORDER
        
        cr = r['capture_ratio']
        if cr is not None and cr != float('inf') and not (isinstance(cr, float) and np.isinf(cr)):
            c7 = ws.cell(row=row, column=7, value=cr)
            c7.number_format = '0.0%'
        else:
            c7 = ws.cell(row=row, column=7, value=r['scenario'])
        c7.border = THIN_BORDER
        
        ws.cell(row=row, column=8, value=r['scenario']).border = THIN_BORDER
        ws.cell(row=row, column=9, value=r['trade_count']).border = THIN_BORDER
        
        c10 = ws.cell(row=row, column=10, value=r['win_rate']/100 if r['win_rate'] else 0)
        c10.number_format = '0%'
        c10.border = THIN_BORDER
        
        c11 = ws.cell(row=row, column=11, value=r['max_single_gain']/100 if r['max_single_gain'] else 0)
        c11.number_format = '+0.0%;-0.0%'
        c11.border = THIN_BORDER
        
        passed = r['passed']
        c12 = ws.cell(row=row, column=12, value='✓' if passed else '✗')
        c12.fill = PASS_FILL if passed else FAIL_FILL
        c12.font = Font(bold=True, color='006100' if passed else '9C0006')
        c12.alignment = Alignment(horizontal='center')
        c12.border = THIN_BORDER
        
        row += 1
    
    # 列宽
    widths = [12, 8, 8, 10, 10, 10, 10, 10, 8, 8, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_failure_analysis_sheet(wb, df):
    """创建失败案例分析页"""
    valid = df[df['status'] == 'success']
    failed = valid[~valid['passed']].sort_values('strategy_return_pct')
    
    if len(failed) == 0:
        return
    
    ws = wb.create_sheet(title='失败案例分析')
    ws.sheet_properties.tabColor = '9C0006'
    
    ws['A1'] = '未达标案例分析'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:J1')
    
    ws['A2'] = '这些标的的捕获率 < 60%，需要重点分析失败原因'
    ws['A2'].font = Font(size=9, color='666666')
    
    row = 4
    headers = ['板块', '代码', '名称', '股价涨幅', '策略收益', '捕获率', '场景', '交易数', '最大亏损', '可能原因']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    
    for _, r in failed.iterrows():
        ws.cell(row=row, column=1, value=r['board']).border = THIN_BORDER
        ws.cell(row=row, column=2, value=r['ts_code']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=r['name']).border = THIN_BORDER
        
        c4 = ws.cell(row=row, column=4, value=r['price_change_pct']/100)
        c4.number_format = '+0.0%;-0.0%'
        c4.border = THIN_BORDER
        
        c5 = ws.cell(row=row, column=5, value=r['strategy_return_pct']/100)
        c5.number_format = '+0.0%;-0.0%'
        c5.border = THIN_BORDER
        
        cr = r['capture_ratio']
        if cr is not None and not (isinstance(cr, float) and np.isinf(cr)):
            c6 = ws.cell(row=row, column=6, value=cr)
            c6.number_format = '0.0%'
        else:
            c6 = ws.cell(row=row, column=6, value='N/A')
        c6.border = THIN_BORDER
        
        ws.cell(row=row, column=7, value=r['scenario']).border = THIN_BORDER
        ws.cell(row=row, column=8, value=r['trade_count']).border = THIN_BORDER
        
        c9 = ws.cell(row=row, column=9, value=r['max_single_loss']/100 if r['max_single_loss'] else 0)
        c9.number_format = '+0.0%;-0.0%'
        c9.border = THIN_BORDER
        
        # 自动诊断可能原因
        reason = ''
        if r['scenario'] == '上涨+亏损':
            if r['trade_count'] <= 1:
                reason = '信号少/入场时机差'
            else:
                reason = '频繁止损，未抓住趋势'
        elif r['scenario'] == '下跌+亏损':
            if abs(r['strategy_return_pct']) > abs(r['price_change_pct']):
                reason = '亏损超过股价跌幅，风控不足'
            else:
                reason = '下跌环境，亏损可控'
        elif r['scenario'] == '上涨+盈利':
            reason = f'捕获率仅{cr:.0%}，止盈过早' if cr and not np.isinf(cr) else '捕获不足'
        
        ws.cell(row=row, column=10, value=reason).border = THIN_BORDER
        
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = FAIL_FILL
        
        row += 1
    
    widths = [8, 12, 8, 10, 10, 10, 10, 8, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_excel_report(csv_path: str = None):
    """主函数：生成Excel报告"""
    if csv_path is None:
        csv_path = OUTPUT_DIR / 'all_boards_backtest_detail.csv'
    
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    print(f"读取 {len(df)} 条回测记录")
    
    wb = Workbook()
    
    # 1. 总览页
    create_summary_sheet(wb, df)
    
    # 2. 各板块明细
    colors = {'创业板': '00B050', '科创板': '2E75B6', '主板': 'FFC000'}
    for board, color in colors.items():
        create_detail_sheet(wb, df, board, color)
    
    # 3. 失败案例分析
    create_failure_analysis_sheet(wb, df)
    
    # 保存
    output_path = OUTPUT_DIR / f'回测报告_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(output_path)
    print(f"报告已保存: {output_path}")
    
    return output_path


if __name__ == '__main__':
    generate_excel_report()
